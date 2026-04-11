"""Запуск скрипта добавления столбца Number"""
import os
from openpyxl import load_workbook

# Определяем путь к папке newtest (используем абсолютный путь)
script_dir = os.path.dirname(os.path.abspath(__file__))
newtest_dir = os.path.join(script_dir, "newtest")

print(f"Директория скрипта: {script_dir}")
print(f"Директория newtest: {newtest_dir}")

# Файлы для обработки
files = [
    os.path.join(newtest_dir, "Опрос для родителей  (Ответы).xlsx"),
    os.path.join(newtest_dir, "Опрос ученика (Ответы).xlsx")
]

print("Начало обработки XLSX файлов...")
print(f"Проверка существования директории newtest: {os.path.exists(newtest_dir)}")

for file_path in files:
    if not os.path.exists(file_path):
        print(f"  ✗ Файл не найден: {os.path.basename(file_path)}")
        continue
    
    print(f"\nОбработка файла: {os.path.basename(file_path)}")
    
    try:
        # Загружаем файл
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Определяем количество заполненных строк (исключая заголовок)
        filled_rows = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is not None:
                filled_rows += 1
        
        print(f"  Найдено заполненных строк: {filled_rows}")
        
        # Вставляем новый столбец между первым и вторым (после колонки A, перед колонкой B)
        ws.insert_cols(2)
        
        # Записываем заголовок в новую колонку B
        ws.cell(row=1, column=2).value = "Number"
        
        # Заполняем столбец числами от 1 до 50 для каждой заполненной строки
        for i in range(1, min(filled_rows + 1, 51)):  # От 1 до 50 или до количества заполненных строк
            ws.cell(row=i + 1, column=2).value = i
        
        # Сохраняем файл
        wb.save(file_path)
        print(f"  ✓ Столбец 'Number' добавлен и заполнен")
        
    except Exception as e:
        print(f"  ✗ Ошибка при обработке: {e}")
        import traceback
        traceback.print_exc()

print("\nОбработка завершена!")

