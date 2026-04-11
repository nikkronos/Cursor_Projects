"""Скрипт для добавления столбца 'Number' между первыми двумя столбцами в XLSX файлах"""
import os
from openpyxl import load_workbook

def add_number_column(file_path):
    """Добавляет столбец 'Number' между первыми двумя столбцами и заполняет его числами от 1 до 50"""
    print(f"\nОбработка файла: {file_path}")
    
    # Загружаем файл
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Определяем количество заполненных строк (исключая заголовок)
    max_row = ws.max_row
    filled_rows = 0
    
    # Подсчитываем заполненные строки (проверяем первую колонку)
    for row in range(2, max_row + 1):
        if ws.cell(row=row, column=1).value is not None:
            filled_rows += 1
    
    print(f"  Найдено заполненных строк: {filled_rows}")
    
    # Вставляем новый столбец между первым и вторым (после колонки A, перед колонкой B)
    ws.insert_cols(2)
    
    # Записываем заголовок в новую колонку B
    ws.cell(row=1, column=2).value = "Number"
    
    # Заполняем столбец числами от 1 до 50 для каждой заполненной строки
    for i in range(1, min(filled_rows + 1, 51)):  # От 1 до 50 или до количества заполненных строк
        ws.cell(row=i + 1, column=2).value = i
    
    # Сохраняем файл
    wb.save(file_path)
    print(f"  ✓ Столбец 'Number' добавлен и заполнен")
    
    return filled_rows

if __name__ == "__main__":
    # Определяем путь к папке со скриптом (используем абсолютный путь)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Файлы для обработки
    files = [
        os.path.join(script_dir, "Опрос для родителей  (Ответы).xlsx"),
        os.path.join(script_dir, "Опрос ученика (Ответы).xlsx")
    ]
    
    print("Начало обработки XLSX файлов...")
    print(f"Рабочая директория: {script_dir}")
    
    for file_path in files:
        file_name = os.path.basename(file_path)
        if os.path.exists(file_path):
            try:
                print(f"\nОбработка файла: {file_name}")
                add_number_column(file_path)
            except Exception as e:
                print(f"  ✗ Ошибка при обработке {file_name}: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"  ✗ Файл не найден: {file_name}")
            print(f"    Полный путь: {file_path}")
    
    print("\nОбработка завершена!")

