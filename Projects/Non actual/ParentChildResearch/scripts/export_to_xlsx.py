"""
Скрипт для экспорта данных в XLSX формат с правильным форматированием
Ответы начинаются с колонки F (без отступов)
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def load_data(parents_file, students_file, pairs_file):
    """Загрузка данных"""
    print("Загрузка данных...")
    
    parents_df = pd.read_csv(parents_file, encoding='utf-8-sig')
    students_df = pd.read_csv(students_file, encoding='utf-8-sig')
    pairs_df = pd.read_csv(pairs_file, encoding='utf-8-sig')
    
    print(f"  Загружено родителей: {len(parents_df)}")
    print(f"  Загружено учеников: {len(students_df)}")
    print(f"  Загружено пар: {len(pairs_df)}")
    
    return parents_df, students_df, pairs_df

def export_to_xlsx(parents_df, students_df, pairs_df, output_file, sheet_name="Data"):
    """
    Экспорт данных в XLSX с правильным форматированием
    Ответы начинаются с колонки F (колонки A-E для метаданных)
    """
    print(f"\nЭкспорт данных в XLSX: {output_file}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Заголовки для родителей
    # Колонки A-E: метаданные
    # Колонка F и далее: ответы на 60 вопросов
    
    # Заголовок для родителей
    parent_headers = ['ID пары', 'Имя родителя', 'Имя ребенка', 'Возраст', 'Дата']
    for i in range(60):
        parent_headers.append(f'Q{i+1}')
    
    # Заголовок для учеников
    student_headers = ['ID пары', 'Имя ребенка', 'Возраст', 'Класс', 'Дата']
    for i in range(60):
        student_headers.append(f'Q{i+1}')
    for i in range(42):
        student_headers.append(f'Choice{i+1}')
    for i in range(20):
        student_headers.append(f'ProfQ{i+1}')
    
    # Записываем данные родителей
    row = 1
    col = 1
    
    # Заголовки родителей
    for header in parent_headers:
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        col += 1
    
    row += 1
    
    # Данные родителей
    pair_num = 0
    for pair_idx, pair in pairs_df.iterrows():
        pair_num += 1
        parent_idx = pair['parent_index']
        parent_row = parents_df.iloc[parent_idx]
        
        col = 1
        # Метаданные (A-E)
        ws.cell(row=row, column=col).value = pair_num
        col += 1
        ws.cell(row=row, column=col).value = pair['parent_name']
        col += 1
        ws.cell(row=row, column=col).value = pair['child_name']
        col += 1
        ws.cell(row=row, column=col).value = parent_row.get('  Возраст ребенка ', '')
        col += 1
        ws.cell(row=row, column=col).value = parent_row.get('Отметка времени', '')
        col += 1
        
        # Ответы на вопросы (F и далее)
        for q in range(60):
            q_col_name = parents_df.columns[5 + q] if 5 + q < len(parents_df.columns) else f'Q{q+1}'
            answer = parent_row.get(q_col_name, '')
            # Конвертируем в число, если возможно
            try:
                answer = int(float(answer)) if answer != '' else ''
            except (ValueError, TypeError):
                pass
            ws.cell(row=row, column=col).value = answer
            col += 1
        
        row += 1
    
    # Добавляем пустую строку
    row += 1
    
    # Записываем данные учеников
    # Заголовки учеников
    col = 1
    for header in student_headers:
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        col += 1
    
    row += 1
    
    # Данные учеников
    pair_num = 0
    for pair_idx, pair in pairs_df.iterrows():
        pair_num += 1
        student_idx = pair['student_index']
        student_row = students_df.iloc[student_idx]
        
        col = 1
        # Метаданные (A-E)
        ws.cell(row=row, column=col).value = pair_num
        col += 1
        ws.cell(row=row, column=col).value = pair['student_name']
        col += 1
        ws.cell(row=row, column=col).value = student_row.get('Возраст ', '')
        col += 1
        ws.cell(row=row, column=col).value = student_row.get('Класс ', '')
        col += 1
        ws.cell(row=row, column=col).value = student_row.get('Отметка времени', '')
        col += 1
        
        # Ответы на вопросы 1-60 (F и далее)
        for q in range(60):
            q_col_name = students_df.columns[4 + q] if 4 + q < len(students_df.columns) else f'Q{q+1}'
            answer = student_row.get(q_col_name, '')
            # Конвертируем в число, если возможно
            try:
                answer = int(float(answer)) if answer != '' else ''
            except (ValueError, TypeError):
                pass
            ws.cell(row=row, column=col).value = answer
            col += 1
        
        # Ответы на вопросы "Что тебе ближе?" (42 вопроса)
        for q in range(42):
            q_col_name = students_df.columns[64 + q] if 64 + q < len(students_df.columns) else f'Choice{q+1}'
            answer = student_row.get(q_col_name, '')
            # Конвертируем в число, если возможно
            try:
                answer = int(float(answer)) if answer != '' else ''
            except (ValueError, TypeError):
                pass
            ws.cell(row=row, column=col).value = answer
            col += 1
        
        # Ответы на вопросы о профессиях (20 вопросов)
        # В данных могут быть колонки с названиями профессий, которые нужно пропустить
        # Собираем все числовые значения из колонок после "Что тебе ближе?"
        prof_start_col = 106  # Начало колонок с вопросами о профессиях
        numeric_ratings = []
        
        # Собираем все числовые значения от начала колонок профессий до конца
        for col_idx in range(prof_start_col, len(students_df.columns)):
            col_name = students_df.columns[col_idx]
            answer = student_row.get(col_name, '')
            
            # Пытаемся конвертировать в число
            try:
                numeric_answer = int(float(answer))
                # Если это число от 1 до 5, это рейтинг
                if 1 <= numeric_answer <= 5:
                    numeric_ratings.append(numeric_answer)
                elif answer == '':
                    numeric_ratings.append('')
            except (ValueError, TypeError):
                # Если не число, пропускаем (это название профессии)
                continue
        
        # Берем последние 20 значений (рейтинги профессий идут после названий)
        prof_ratings = numeric_ratings[-20:] if len(numeric_ratings) >= 20 else numeric_ratings
        
        # Записываем рейтинги
        for rating in prof_ratings:
            ws.cell(row=row, column=col).value = rating
            col += 1
        
        # Если не хватает рейтингов, заполняем пустыми значениями
        while len(prof_ratings) < 20:
            ws.cell(row=row, column=col).value = ''
            col += 1
            prof_ratings.append('')
        
        row += 1
    
    # Сохраняем файл
    wb.save(output_file)
    print(f"  Файл сохранен: {output_file}")

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.join(base_dir, '..')
    
    # Пути к файлам
    parents_file = os.path.join(project_root, 'data', 'raw', 'parents_survey.csv')
    students_file = os.path.join(project_root, 'data', 'raw', 'students_survey_converted.csv')
    pairs_file = os.path.join(project_root, 'data', 'raw', 'pairs_mapping.csv')
    
    # Проверяем существование файлов
    if not os.path.exists(parents_file):
        print(f"Ошибка: файл {parents_file} не найден")
        exit(1)
    
    if not os.path.exists(students_file):
        print(f"Ошибка: файл {students_file} не найден")
        exit(1)
    
    if not os.path.exists(pairs_file):
        print(f"Ошибка: файл {pairs_file} не найден")
        exit(1)
    
    # Загружаем данные
    parents_df, students_df, pairs_df = load_data(parents_file, students_file, pairs_file)
    
    # Экспортируем существующие пары
    output_dir = os.path.join(project_root, 'data', 'generated')
    os.makedirs(output_dir, exist_ok=True)
    
    existing_output = os.path.join(output_dir, 'existing_pairs.xlsx')
    export_to_xlsx(parents_df, students_df, pairs_df, existing_output, "Existing Pairs")
    
    # Если есть сгенерированные данные, экспортируем все 50 пар
    new_parents_file = os.path.join(project_root, 'data', 'generated', 'new_parents.csv')
    new_students_file = os.path.join(project_root, 'data', 'generated', 'new_students.csv')
    
    if os.path.exists(new_parents_file) and os.path.exists(new_students_file):
        print("\nЗагрузка сгенерированных данных...")
        new_parents_df = pd.read_csv(new_parents_file, encoding='utf-8-sig')
        new_students_df = pd.read_csv(new_students_file, encoding='utf-8-sig')
        
        # Объединяем существующие и новые данные
        all_parents_df = pd.concat([parents_df, new_parents_df], ignore_index=True)
        all_students_df = pd.concat([students_df, new_students_df], ignore_index=True)
        
        # Создаем пары для всех данных
        all_pairs = []
        for i in range(len(pairs_df)):
            all_pairs.append(pairs_df.iloc[i].to_dict())
        
        # Добавляем пары для новых данных
        for i in range(len(new_parents_df)):
            all_pairs.append({
                'parent_index': len(parents_df) + i,
                'student_index': len(students_df) + i,
                'parent_name': new_parents_df.iloc[i].get('Фамилия, имя родителя ', ''),
                'child_name': new_parents_df.iloc[i].get('  Фамилия и имя ребенка  ', ''),
                'student_name': new_students_df.iloc[i].get('Фамилия и имя ', ''),
                'match_score': 0
            })
        
        all_pairs_df = pd.DataFrame(all_pairs)
        
        # Экспортируем все 50 пар
        all_output = os.path.join(output_dir, 'all_50_pairs.xlsx')
        export_to_xlsx(all_parents_df, all_students_df, all_pairs_df, all_output, "All 50 Pairs")
    
    print("\nЭкспорт завершен!")








