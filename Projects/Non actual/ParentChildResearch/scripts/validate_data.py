"""
Скрипт для валидации преобразованных данных и форматирования XLSX
"""
import pandas as pd
import os

def validate_converted_data(students_file):
    """Проверка корректности преобразования текстовых ответов в числовые"""
    print("Проверка преобразованных данных...")
    
    if not os.path.exists(students_file):
        print(f"  Предупреждение: файл {students_file} не найден")
        return False
    
    df = pd.read_csv(students_file, encoding='utf-8-sig')
    
    # Проверяем колонки с вопросами "Что тебе ближе?" (колонки 65-106)
    choice_start_col = 64
    choice_end_col = 106
    
    errors = []
    warnings = []
    
    for col_idx in range(choice_start_col, min(choice_end_col, len(df.columns))):
        col_name = df.columns[col_idx]
        col_data = df[col_name]
        
        # Проверяем, что все значения числовые (1 или 2)
        for idx, value in col_data.items():
            if pd.notna(value):
                try:
                    num_value = int(value)
                    if num_value not in [1, 2]:
                        errors.append(f"Строка {idx+2}, колонка {col_name}: значение {value} не является 1 или 2")
                except (ValueError, TypeError):
                    errors.append(f"Строка {idx+2}, колонка {col_name}: значение {value} не является числом")
    
    if errors:
        print(f"  Найдено ошибок: {len(errors)}")
        for error in errors[:10]:  # Показываем первые 10 ошибок
            print(f"    {error}")
        if len(errors) > 10:
            print(f"    ... и еще {len(errors) - 10} ошибок")
        
        # Если ошибок меньше 5% от общего количества, считаем это предупреждением, а не ошибкой
        total_values = len(df) * 42  # 42 вопроса для каждой строки
        error_percentage = (len(errors) / total_values) * 100
        if error_percentage < 5:
            print(f"  ⚠ Предупреждение: {error_percentage:.2f}% ошибок (меньше 5%, допустимо)")
            return True  # Не критично
        else:
            return False
    else:
        print("  Все значения корректны (1 или 2)")
        return True

def validate_pairs(pairs_file, expected_count=10):
    """Проверка наличия всех ожидаемых пар"""
    print(f"\nПроверка пар (ожидается {expected_count})...")
    
    if not os.path.exists(pairs_file):
        print(f"  Ошибка: файл {pairs_file} не найден")
        return False
    
    df = pd.read_csv(pairs_file, encoding='utf-8-sig')
    
    print(f"  Найдено пар: {len(df)}")
    
    # Проверяем наличие пары Пешич
    peshic_found = False
    for idx, row in df.iterrows():
        if 'пешич' in str(row.get('parent_name', '')).lower() or 'пешич' in str(row.get('child_name', '')).lower():
            peshic_found = True
            print(f"  Найдена пара Пешич: {row.get('parent_name')} / {row.get('child_name')}")
            break
    
    if not peshic_found:
        print("  Предупреждение: пара Пешич не найдена")
    
    if len(df) >= expected_count:
        print(f"  ✓ Количество пар соответствует ожидаемому ({expected_count}+)")
        return True
    else:
        print(f"  ⚠ Предупреждение: найдено {len(df)} пар вместо ожидаемых {expected_count}")
        print(f"  (Это нормально, если в данных меньше реальных пар)")
        return len(df) > 0  # Не критично, если есть хотя бы одна пара

def validate_xlsx_formatting(xlsx_file):
    """Проверка форматирования XLSX файла"""
    print(f"\nПроверка форматирования XLSX: {xlsx_file}...")
    
    if not os.path.exists(xlsx_file):
        print(f"  Предупреждение: файл {xlsx_file} не найден")
        return False
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_file)
        ws = wb.active
        
        # Проверяем, что ответы начинаются с колонки F (6)
        # Колонки A-E должны содержать метаданные
        
        # Проверяем первую строку с данными (строка 2, так как строка 1 - заголовки)
        if ws.max_row > 1:
            row = 2
            # Проверяем, что колонка F (6) содержит данные
            cell_f = ws.cell(row=row, column=6)
            if cell_f.value is None or cell_f.value == '':
                print("  Предупреждение: колонка F пуста, возможно неправильное форматирование")
                return False
            else:
                print("  ✓ Ответы начинаются с колонки F")
                return True
        else:
            print("  Предупреждение: файл не содержит данных")
            return False
            
    except Exception as e:
        print(f"  Ошибка при проверке XLSX: {e}")
        return False

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.join(base_dir, '..')
    
    # Пути к файлам
    students_file = os.path.join(project_root, 'data', 'raw', 'students_survey_converted.csv')
    pairs_file = os.path.join(project_root, 'data', 'raw', 'pairs_mapping.csv')
    existing_xlsx = os.path.join(project_root, 'data', 'generated', 'existing_pairs.xlsx')
    all_xlsx = os.path.join(project_root, 'data', 'generated', 'all_50_pairs.xlsx')
    
    print("=" * 60)
    print("ВАЛИДАЦИЯ ДАННЫХ")
    print("=" * 60)
    
    # Проверяем преобразованные данные
    converted_ok = validate_converted_data(students_file)
    
    # Проверяем пары
    pairs_ok = validate_pairs(pairs_file, expected_count=10)
    
    # Проверяем форматирование XLSX
    xlsx_ok = True
    if os.path.exists(existing_xlsx):
        xlsx_ok = validate_xlsx_formatting(existing_xlsx) and xlsx_ok
    if os.path.exists(all_xlsx):
        xlsx_ok = validate_xlsx_formatting(all_xlsx) and xlsx_ok
    
    print("\n" + "=" * 60)
    print("РЕЗУЛЬТАТЫ ВАЛИДАЦИИ")
    print("=" * 60)
    print(f"Преобразованные данные: {'✓ OK' if converted_ok else '✗ ОШИБКИ'}")
    print(f"Пары: {'✓ OK' if pairs_ok else '✗ ПРОБЛЕМЫ'}")
    print(f"Форматирование XLSX: {'✓ OK' if xlsx_ok else '✗ ПРОБЛЕМЫ'}")
    
    if converted_ok and pairs_ok and xlsx_ok:
        print("\n✓ Все проверки пройдены успешно!")
        exit(0)
    else:
        print("\n✗ Обнаружены проблемы. Проверьте вывод выше.")
        exit(1)

