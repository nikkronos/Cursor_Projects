"""Создание XLSX файла с данными"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font

# Определяем пути
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.join(script_dir, '..')
# Меняем рабочую директорию на корень проекта
os.chdir(project_root)

parents_file = os.path.join(project_root, 'data', 'raw', 'parents_survey.csv')
students_file = os.path.join(project_root, 'data', 'raw', 'students_survey_converted.csv')
pairs_file = os.path.join(project_root, 'data', 'raw', 'pairs_mapping.csv')
output_dir = os.path.join(project_root, 'data', 'generated')
os.makedirs(output_dir, exist_ok=True)

print("Загрузка данных...")
parents_df = pd.read_csv(parents_file, encoding='utf-8-sig')
students_df = pd.read_csv(students_file, encoding='utf-8-sig')
pairs_df = pd.read_csv(pairs_file, encoding='utf-8-sig')

print(f"  Загружено родителей: {len(parents_df)}")
print(f"  Загружено учеников: {len(students_df)}")
print(f"  Загружено пар: {len(pairs_df)}")

# Создаем XLSX
output_file = os.path.join(output_dir, 'existing_pairs.xlsx')
print(f"\nЭкспорт данных в XLSX: {output_file}")

wb = Workbook()
ws = wb.active
ws.title = "Existing Pairs"

# Заголовки для родителей
parent_headers = ['ID пары', 'Имя родителя', 'Имя ребенка', 'Возраст', 'Дата']
for i in range(60):
    parent_headers.append(f'Q{i+1}')

# Заголовки для учеников
student_headers = ['ID пары', 'Имя ребенка', 'Возраст', 'Класс', 'Дата']
for i in range(60):
    student_headers.append(f'Q{i+1}')
for i in range(42):
    student_headers.append(f'Choice{i+1}')
for i in range(20):
    student_headers.append(f'ProfQ{i+1}')

# Записываем данные родителей
row = 1
col = 1

# Заголовки родителей
for header in parent_headers:
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    col += 1

row += 1

# Данные родителей
for pair_idx, pair in pairs_df.iterrows():
    parent_idx = pair['parent_index']
    parent_row = parents_df.iloc[parent_idx]
    
    col = 1
    # Метаданные (A-E)
    ws.cell(row=row, column=col).value = pair_idx + 1
    col += 1
    ws.cell(row=row, column=col).value = pair['parent_name']
    col += 1
    ws.cell(row=row, column=col).value = pair['child_name']
    col += 1
    ws.cell(row=row, column=col).value = parent_row.get('  Возраст ребенка ', '')
    col += 1
    ws.cell(row=row, column=col).value = parent_row.get('Отметка времени', '')
    col += 1
    
    # Ответы на вопросы (F и далее)
    for q in range(60):
        q_col_name = parents_df.columns[5 + q] if 5 + q < len(parents_df.columns) else f'Q{q+1}'
        answer = parent_row.get(q_col_name, '')
        ws.cell(row=row, column=col).value = answer
        col += 1
    
    row += 1

# Добавляем пустую строку
row += 1

# Записываем данные учеников
# Заголовки учеников
col = 1
for header in student_headers:
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    col += 1

row += 1

# Данные учеников
for pair_idx, pair in pairs_df.iterrows():
    student_idx = pair['student_index']
    student_row = students_df.iloc[student_idx]
    
    col = 1
    # Метаданные (A-E)
    ws.cell(row=row, column=col).value = pair_idx + 1
    col += 1
    ws.cell(row=row, column=col).value = pair['student_name']
    col += 1
    ws.cell(row=row, column=col).value = student_row.get('Возраст ', '')
    col += 1
    ws.cell(row=row, column=col).value = student_row.get('Класс ', '')
    col += 1
    ws.cell(row=row, column=col).value = student_row.get('Отметка времени', '')
    col += 1
    
    # Ответы на вопросы 1-60 (F и далее)
    for q in range(60):
        q_col_name = students_df.columns[4 + q] if 4 + q < len(students_df.columns) else f'Q{q+1}'
        answer = student_row.get(q_col_name, '')
        ws.cell(row=row, column=col).value = answer
        col += 1
    
    # Ответы на вопросы "Что тебе ближе?" (42 вопроса)
    for q in range(42):
        q_col_name = students_df.columns[64 + q] if 64 + q < len(students_df.columns) else f'Choice{q+1}'
        answer = student_row.get(q_col_name, '')
        ws.cell(row=row, column=col).value = answer
        col += 1
    
    # Ответы на вопросы о профессиях (20 вопросов)
    for q in range(20):
        q_col_name = students_df.columns[106 + q] if 106 + q < len(students_df.columns) else f'ProfQ{q+1}'
        answer = student_row.get(q_col_name, '')
        ws.cell(row=row, column=col).value = answer
        col += 1
    
    row += 1

# Сохраняем файл
wb.save(output_file)
print(f"  Файл сохранен: {output_file}")
print("\nЭкспорт завершен!")





